import asyncio
import sys
import random
from pathlib import Path
from playwright.async_api import async_playwright
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).parent
DOWNLOADS_DIR = PROJECT_ROOT / "downloads"
DOWNLOADS_DIR.mkdir(exist_ok=True)
EXCEL_FILE = PROJECT_ROOT / "serials.xlsx"

# List of user-agents to rotate
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
]

def load_serials_from_excel():
    # Load all serials from Excel
    all_serials = []
    try:
        wb = load_workbook(EXCEL_FILE)
        sheet = wb["Серийники"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[1]:
                all_serials.append(row[1].strip().upper())
        print(f"✓ Loaded {len(all_serials)} serials from Excel.")
    except Exception as e:
        print(f"✗ Error loading Excel file: {e}")
        sys.exit(1)
    
    # Get list of already downloaded serials
    downloaded_serials = set()
    for file_path in DOWNLOADS_DIR.glob("PartsExport_Serial-*.xlsx"):
        filename = file_path.name
        parts = filename.split('_')
        if len(parts) >= 2 and parts[1].startswith('Serial-'):
            serial = parts[1][7:].upper()
            downloaded_serials.add(serial)
    
    print(f"✓ Found {len(downloaded_serials)} already downloaded serials.")
    
    # Filter out downloaded ones
    remaining_serials = [s for s in all_serials if s not in downloaded_serials]
    skipped = len(all_serials) - len(remaining_serials)
    print(f"✓ Skipping {skipped} downloaded serials. Processing {len(remaining_serials)} remaining.")
    
    return remaining_serials

async def download_lenovo_parts(serial: str, context, index: int, total: int) -> str:
    user_agent = random.choice(USER_AGENTS)
    page = await context.new_page()
    await page.set_extra_http_headers({"User-Agent": user_agent})
    
    try:
        url = f"https://datacentersupport.lenovo.com/lv/ru/products/servers/thinksystem/sr665/7d2v/7d2vcto1ww/{serial.lower()}/parts/display/as-built"
        print(f"Processing {index}/{total}: {serial}")
        
        response = await page.goto(url, wait_until="domcontentloaded", timeout=60000)
        if response and response.url.endswith("pagenotfound"):
            print(f"✗ Redirected to not found for {serial}.")
            return None
        
        await asyncio.sleep(2)
        
        button = page.locator('div.download-style')
        await button.first.wait_for(state="visible", timeout=10000)
        
        await button.first.scroll_into_view_if_needed()
        await asyncio.sleep(1)
        
        await button.first.hover()
        await asyncio.sleep(0.5)
        
        bounding_box = await button.first.bounding_box()
        if bounding_box:
            center_x = bounding_box['x'] + bounding_box['width'] / 2
            center_y = bounding_box['y'] + bounding_box['height'] / 2
            await page.mouse.move(center_x, center_y)
            await asyncio.sleep(0.5)
        else:
            return None
        
        async with page.expect_download(timeout=30000) as download_info:
            await page.mouse.click(center_x, center_y, button="left", delay=100)
        
        download = await download_info.value
        filename = download.suggested_filename or f"{serial}_parts.xlsx"
        file_path = DOWNLOADS_DIR / filename
        
        await download.save_as(file_path)
        print(f"✓ Downloaded: {filename}")
        
        return filename
        
    except Exception as e:
        print(f"✗ Error for {serial}: {e}")
        return None
        
    finally:
        await page.close()

async def main():
    serials = load_serials_from_excel()
    if not serials:
        return
    
    total = len(serials)
    async with async_playwright() as p:
        user_data_dir = PROJECT_ROOT / "lenovo_cookies"
        context = await p.chromium.launch_persistent_context(
            user_data_dir=user_data_dir,
            headless=False,  # Visible browser
            accept_downloads=True
        )
        
        successes = 0
        failures = 0
        for i, serial in enumerate(serials, 1):
            result = await download_lenovo_parts(serial, context, i, total)
            if result:
                successes += 1
            else:
                failures += 1
            
            # Delay between serials
            if i < total:
                print("Waiting 5s before next serial...")
                await asyncio.sleep(5)

if __name__ == "__main__":
    asyncio.run(main())