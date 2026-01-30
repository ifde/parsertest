import csv
from pathlib import Path
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).parent
EXCEL_FILE = PROJECT_ROOT / "output.xlsx"  # The existing mega file with parts data
MODELS_CSV = PROJECT_ROOT / "models.csv"  # CSV file for models

def load_serials_from_excel():
    """Load all serials from Excel sheet 'Серийники'."""
    all_serials = []
    try:
        wb = load_workbook(EXCEL_FILE)
        sheet = wb["Серийники"]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[1]:  # Assuming serials are in column B (index 1)
                all_serials.append(str(row[1]).strip().upper())
        print(f"✓ Loaded {len(all_serials)} serials from Excel.")
    except Exception as e:
        print(f"✗ Error loading Excel file: {e}")
        return []
    return all_serials

def load_models_serials():
    """Load serials that have models from CSV."""
    models_serials = set()
    try:
        with open(MODELS_CSV, 'r', newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                serial = row.get('Serial', '').strip().upper()
                if serial:
                    models_serials.add(serial)
        print(f"✓ Loaded {len(models_serials)} serials with models from {MODELS_CSV}.")
    except Exception as e:
        print(f"✗ Error loading models CSV: {e}")
        return set()
    return models_serials

def main():
    all_serials = load_serials_from_excel()
    models_serials = load_models_serials()
    
    missing_serials = [s for s in all_serials if s not in models_serials]
    
    print(f"\nSerials in Excel but not in models.csv ({len(missing_serials)} total):")
    for serial in missing_serials:
        print(serial)

if __name__ == "__main__":
    main()