import os
import openpyxl
from openpyxl import load_workbook

# Paths
exports_dir = 'downloads'  # Directory with separate PartsExport_*.xlsx files
mega_file = 'output.xlsx'  # The existing mega file to append to

# Load the mega workbook
wb = load_workbook(mega_file)
sheet = wb['Состав']  # Select the sheet named "Состав"

# Process each separate file
for filename in os.listdir(exports_dir):
    if filename.startswith('PartsExport_Serial-') and filename.endswith('.xlsx'):
        file_path = os.path.join(exports_dir, filename)
        
        # Extract serial number from filename (e.g., '06fm735' from 'PartsExport_Serial-06fm735_...')
        serial = filename.split('Serial-')[1].split('_')[0]
        
        # Load separate workbook
        wb_sep = load_workbook(file_path)
        sheet_sep = wb_sep.active
        
        # Skip header row, process data rows
        for row in sheet_sep.iter_rows(min_row=2, values_only=True):
            # Unpack 7 columns: Description, Commodity Type, Part Number, Installed Qty, MFG Part Number, (empty), Customer Serviceable
            if len(row) >= 7:
                desc, comm_type, part_num, qty, mfg_part, empty, cust_serv = row[:7]
                
                # Append serial as the 8th column
                row_list = [desc, comm_type, part_num, qty, mfg_part, empty, cust_serv, serial]
                
                # Append to mega sheet
                sheet.append(row_list)

# Save the updated mega file
wb.save(mega_file)
print("All files combined into", mega_file, "sheet 'Состав'")