import csv
from pathlib import Path
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).parent
EXCEL_FILE = PROJECT_ROOT / "output.xlsx"  # The existing mega file with parts data
MODELS_CSV = PROJECT_ROOT / "models.csv"  # CSV file for models

def load_models_from_csv():
    """Load models from CSV."""
    models = {}
    try:
        with open(MODELS_CSV, 'r', newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                serial = row.get('Serial', '').strip().upper()
                model = row.get('Model', 'N/A')
                if serial:
                    models[serial] = model
        print(f"✓ Loaded {len(models)} models from {MODELS_CSV}.")
    except Exception as e:
        print(f"✗ Error loading models from CSV: {e}")
        return {}
    return models

def update_excel_with_models(models):
    """Update the Excel file with models."""
    try:
        wb = load_workbook(EXCEL_FILE)
        sheet = wb["Состав"]
        
        # Get header values as a list
        header_row_values = [cell.value for cell in sheet[1]]
        
        # Add header for Model column (after Serial) if not already there
        if "Model" not in header_row_values:
            header_row_values.append("Model")
            sheet.delete_rows(1)
            sheet.insert_rows(1)
            for col, value in enumerate(header_row_values, 1):
                sheet.cell(row=1, column=col, value=value)
        
        # Determine column indices (1-based)
        serial_col_idx = sheet.max_column - 1  # Serial is second last
        model_col_idx = sheet.max_column     # Model is last
        
        # Update rows with model based on serial using iter_rows for speed
        updated_count = 0
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            serial_cell = row[serial_col_idx - 1]  # 0-based index
            serial = str(serial_cell.value or "").strip().upper()
            model = models.get(serial, "N/A")
            model_cell = row[model_col_idx - 1]
            if model_cell.value != model:  # Only update if different
                model_cell.value = model
                updated_count += 1
            print(row)
        
        wb.save(EXCEL_FILE)
        print(f"✓ Excel updated with models in output.xlsx (updated {updated_count} rows)")
    except Exception as e:
        print(f"✗ Error updating Excel: {e}")

def main():
    models = load_models_from_csv()
    if models:
        update_excel_with_models(models)

if __name__ == "__main__":
    main()