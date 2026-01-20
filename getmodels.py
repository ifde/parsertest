import asyncio
import sys
import random
import os
import csv
from pathlib import Path
from playwright.async_api import async_playwright
from openpyxl import load_workbook
import time

PROJECT_ROOT = Path(__file__).parent
EXCEL_FILE = PROJECT_ROOT / "output.xlsx"  # The existing mega file with parts data
MODELS_CSV = PROJECT_ROOT / "models.csv"  # CSV file for models

# List of user-agents to rotate
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
]

def load_serials_from_excel():
    """Load unique serials from the Excel file."""
    unique_serials = set()
    try:
        wb = load_workbook(EXCEL_FILE)
        sheet = wb["Серийники"]
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
            serial = row[-1]  # Last column is serial
            if serial:
                unique_serials.add(str(serial).strip().upper())
        print(f"✓ Loaded {len(unique_serials)} unique serials from Excel.")
    except Exception as e:
        print(f"✗ Error loading Excel file: {e}")
        sys.exit(1)
    return list(unique_serials)

def load_existing_models():
    """Load existing models from CSV if it exists."""
    models = {}
    if MODELS_CSV.exists():
        try:
            with open(MODELS_CSV, 'r', newline='', encoding='utf-8') as csvfile:
                reader = csv.DictReader(csvfile)
                for row in reader:
                    serial = row.get('Serial', '').strip().upper()
                    model = row.get('Model', 'N/A')
                    if serial:
                        models[serial] = model
            print(f"✓ Loaded {len(models)} existing models from {MODELS_CSV}.")
        except Exception as e:
            print(f"✗ Error loading existing models: {e}")
    else:
        print("No existing models.csv found; starting fresh.")
    return models

async def get_model_for_serial(serial: str, context, index: int, total: int) -> str:
    user_agent = random.choice(USER_AGENTS)
    page = await context.new_page()
    await page.set_extra_http_headers({"User-Agent": user_agent})
    
    try:
        # Navigate to the parts page
        base_url = "https://datacentersupport.lenovo.com/lv/ru/products/servers/thinksystem/sr665/7d2v/7d2vcto1ww/parts"
        print(f"Processing {index}/{total}: {serial} - Navigating to parts page")
        
        await page.goto(base_url, wait_until="domcontentloaded", timeout=60000)

        # Handle country modal
        try:
            country_modal = page.locator('#ipdetect_differentCountryModal')
            if await country_modal.is_visible(timeout=5000):
                continue_button = country_modal.locator('.btn_no')
                await continue_button.click()
        except Exception as e:
            print(f"Country modal handling: {e}")
        
        # Accept Evidon cookies
        try:
            accept_button = page.locator('#_evidon-banner-acceptbutton')
            if await accept_button.is_visible(timeout=5000):
                await accept_button.click()
        except Exception as e:
            print(f"Cookie banner handling: {e}")
        
        # Enter serial in the input field
        input_field = page.locator('input.sn-input-sec-nav.typeahead.tt-input')
        await input_field.fill(serial)
        print(f"Entered serial: {serial}")
        
        # Click the search button
        search_button = page.locator('span.sn-title-icon.inputing.icon-l-right.inputmode[role="button"]')
        await search_button.click()
        print("Clicked search")
        
        # Wait for the product name text to appear
        prod_name_locator = page.locator('div.prod-name-text')
        await prod_name_locator.wait_for(state="visible", timeout=30000)
        
        # Scrape the model
        model_text = await prod_name_locator.text_content()
        model = model_text.strip() if model_text else "N/A"
        print(f"✓ Scraped model for {serial}: {model}")
        
        return model
        
    except Exception as e:
        print(f"✗ Error for {serial}: {e}")
        return "N/A"
        
    finally:
        await page.close()

async def main():
    # Load all serials
    all_serials = load_serials_from_excel()
    
    # Load existing models
    models = load_existing_models()
    
    # Filter to only unprocessed serials
    remaining_serials = [s for s in all_serials if s not in models]
    if not remaining_serials:
        print("All serials already processed. Exiting.")
        # return
    
    total = len(remaining_serials)
    print(f"Processing {total} remaining serials (skipped {len(all_serials) - total} already done).")
    start_time = time.time()
    
    async with async_playwright() as p:
        user_data_dir = PROJECT_ROOT / "lenovo_cookies"
        user_data_dir.mkdir(parents=True, exist_ok=True)
        context = await p.chromium.launch_persistent_context(
            user_data_dir=user_data_dir,
            headless=False,  # Real browser
            args=[
                "--window-position=-10000,-10000",  # Off-screen
                "--window-size=1,1",  # Tiny
                "--disable-background-timer-throttling",  # Keep active in background
                "--disable-blink-features=AutomationControlled"
            ]
        )
        
        # Minimize and background the browser window (macOS specific)
        os.system('osascript -e \'tell application "Google Chrome for Testing" to set miniaturized of window 1 to true\'')
        os.system('osascript -e \'tell application "Google Chrome for Testing" to set frontmost of frontmost to false\'')
        print("Browser window minimized and sent to background.")
        
        successes = 0
        failures = 0
        iteration_count = 0  # Counter for restarting context
        for i, serial in enumerate(remaining_serials, 1):
            iteration_count += 1
            
            # Restart browser context every 50 iterations
            if iteration_count % 50 == 0:
                print(f"Restarting browser context after {iteration_count} iterations to prevent JavaScript storage issues.")
                await context.close()
                context = await p.chromium.launch_persistent_context(
                    user_data_dir=user_data_dir,
                    headless=False,
                    args=[
                        "--window-position=-10000,-10000",
                        "--window-size=1,1",
                        "--disable-background-timer-throttling",
                        "--disable-blink-features=AutomationControlled"
                    ]
                )
                # Re-minimize and background the new browser window
                os.system('osascript -e \'tell application "Google Chrome for Testing" to set miniaturized of window 1 to true\'')
                os.system('osascript -e \'tell application "Google Chrome for Testing" to set frontmost of frontmost to false\'')
                print("Browser context restarted and window minimized.")
            
            model = await get_model_for_serial(serial, context, i, total)
            models[serial] = model  # Add to dict
            if model != "N/A":
                successes += 1
            else:
                failures += 1
            
            # Save progress every 10 serials (optional, for safety)
            if i % 10 == 0:
                with open(MODELS_CSV, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerow(["Serial", "Model"])
                    for s, m in models.items():
                        writer.writerow([s, m])
                print(f"Progress saved after {i} serials.")
        
        await context.close()
    
    # Save all models to CSV
    print("Saving all models to CSV...")
    with open(MODELS_CSV, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(["Serial", "Model"])
        for serial, model in models.items():
            writer.writerow([serial, model])
    print(f"✓ All models saved to {MODELS_CSV}")
    
    # Update the Excel file with models
    print("Updating Excel with models...")
    wb = load_workbook(EXCEL_FILE)
    sheet = wb["Состав"]
    
    # Get header values as a list
    header_row_values = [cell.value for cell in sheet[1]]
    
    # Add header for Model column (after Serial) if not already there
    if "Model" not in header_row_values:
        header_row_values.append("Model")
        sheet.delete_rows(1)
        sheet.insert_rows(1)
        for col, value in enumerate(header_row_values, 1):
            sheet.cell(row=1, column=col, value=value)
    
    # Update rows with model based on serial
    for row_idx in range(2, sheet.max_row + 1):
        serial_cell = sheet.cell(row=row_idx, column=sheet.max_column - 1)  # Serial is second last now
        serial = str(serial_cell.value or "").strip().upper()
        model = models.get(serial, "N/A")
        sheet.cell(row=row_idx, column=sheet.max_column, value=model)  # Add to last column
        print(f"Index: {row_idx}")
    
    wb.save(EXCEL_FILE)
    
    total_time = time.time() - start_time
    print(f"Processed {total} serials: {successes} successes, {failures} failures")
    print(f"Total time: {total_time:.2f} seconds")
    print("Excel updated with models in output.xlsx")

if __name__ == "__main__":
    asyncio.run(main())